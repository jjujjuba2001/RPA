import openpyxl as op

wb = op.Workbook()
ws = wb.create_sheet ("지출내역서")

print(ws)
ws = wb["지출내역서"]

title = ["날짜","항목","가격","인원","총 가격(단가*인원)"]
ws.append (title)

data = [['2022-07-01', '다과비', 5000,3],
        ['2022-07-02', '점심 식대', 10000,2],
        ['2022-07-02', '저녁 식대', 13000,1],
        ['2022-07-03', '유류비', 100000,3],
        ['2022-07-04', 'coffee', 15000,5],
        ['2024-07-12', '가나다', 254000,2],
        ['2024-08-23', 'lunch', 18450,5]
        ]


for cell_data in data:
    ws.append (cell_data)

#ws["C8"] = "=sum(C2:C7)"

row_max = ws.max_row
col_max = ws.max_column

print ("로우 최대값 : ", row_max)
print ("컬럼 최대값 : ", col_max)

'''for row_max in range (2, row_max+1):   #내가 만든 곱셈 로직
    print(row_max)
    ws.cell(row_max,col_max).value = "=C"+ str(row_max) + "*" + "D"+ str(row_max)
    row_max -= 1
    print(row_max)'''
    
for row in range(2, row_max+1):     #책에서 나온 곱셈 로직
    ws["E"+str(row)] = "=C" + str(row) + "*" + "D" + str(row)

'''for sum_row in range (row_max+1,col_max+1):   # 합계 구하기 ing 중
    ws.cell(row_max,col_max).value = "=SUM"'''

'''for row in range (2, row_max+1):
    ws["E"+str(row)] = ws["C"+str(row)].value * ws["D"+str(row)].value'''

read_data = []

for low in ws.rows:
    read_data.append(low[4].value)

print (read_data)

wb.save (r"C:\Users\user\Documents\Code\지출내역서.xlsx")
           
wb.close()
                     
