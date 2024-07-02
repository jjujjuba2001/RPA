import openpyxl as op
from openpyxl.styles.fonts import Font
from openpyxl.styles import Border, Side
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill


wb = op.load_workbook(r"C:\Users\user\Documents\Code\sample.xlsx")
ws = wb.active

ws.row_dimensions[5].height = 70            #5행 높이 70 설정
ws.row_dimensions[6].height = 70
ws.column_dimensions["A"].width = 70        #A열 폭 70 설정

ws["A5"].value = "사장님 모올래 자동하"         #A5 셀에 값 입력
ws["A5"].font = Font(size=22, name="궁서체",color="FF0000") #underline="doubleAccounting"
ws["A5"].alignment = Alignment (horizontal="right", vertical= "bottom")
ws["A5"].border = Border (top=Side(border_style="dashDotDot", color="000080"), right=Side(border_style="double"))
ws["A5"].fill = PatternFill (fill_type="solid",fgColor="000080")

ws["A6"].value = "Actually it is hard work more than past 아 몰라"
font_format = Font (color="000080", bold=True, italic=True, size=30) #strike=True           #변수에 폰트 설정 할당
ws["A6"].font = font_format                 # A6 셀에 변수에 할당한 폰트 설정 적용
alig_format = Alignment (horizontal= "center", vertical= "center")
ws["A6"].alignment = alig_format
ws["A6"].fill = PatternFill (fill_type="lightGray", fgColor="FF6600")


wb.save ("sample.xlsx")
wb.close()
