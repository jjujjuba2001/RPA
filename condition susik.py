import openpyxl as op
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment

#wb = op.Workbook()
#ws = wb.create_sheet("성적표")

path = r"C:\Users\user\Documents\RPAStudy\성적표.xlsx"
wb = op.load_workbook(path, data_only=True)
ws = wb["성적표"]

data = [["이름", '국어', '영어', '수학', '평균', '합/불 여부'],
        ['영배', 90, 100, 100],
        ['수원', 100, 55, 98],
        ['비제이', 80, 85, 77],
        ['미자', 100, 99, 80]]

'''for row in data:
    ws.append(row)
    print (row)'''

def pass_condition():
    pass_format = Font(size=15, name='바탕', color='000000FF')
    pass_fill_format = PatternFill(fgColor='FFFF00')

    max_row = ws.max_row



def fail_condition():
    fail_format = Font(size=18, name='굴림', color='00FF0000')
    fail_fill_format = PatternFill(fgColor='000000FF')

def write_result():
    max_row = ws.max_row
    print ("최대 열은 : "+ str(max_row))

    for row_index in range (2, max_row+1):
        print("포문 : " + str(row_index))
        average = ws.cell(row = row_index, column = 5).value

        if average >= 85:
            ws.cell(row = row_index, column = 6).value = "합격"
            ws.cell(row = row_index, column = 6).value = pass_condition()
        else:
            ws.cell(row = row_index, column = 6).value = "불합격"
            ws.cell(row = row_index, column = 6).value = fail_condition()

write_result()
    
wb.save("성적표.xlsx")
wb.close()